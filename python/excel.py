from openpyxl.styles import Font 
import matplotlib.pyplot as plt 
from openpyxl import Workbook
import datetime, openpyxl 
import pandas as pd, os

class Excel:
    
    def __init__(self, file_path):
        self.file_path = file_path 
        if not os.path.exists(file_path):
            self.CreateSpreadsheet()
        self._log('Inicialização da Classe Excel')
        
    def _log(self, message):
        with open('vitalynk.excel.logs.txt', 'a') as log_file:
            log_file.write(f'{datetime.datetime.now()} - {message}\n')

    def CreateSpreadsheet(self):
        wb = Workbook() 
        wb.save(self.file_path)
        self._log('Planilha criada')
    
    def ReadSpreadsheet(self, sheet_name = 'Sheet1'):
        df = pd.read_excel(self.file_path, sheet_name = sheet_name)
        self._log(f"Planilha '{sheet_name}' lida")
        return df 

    def DeleteSpreadsheet(self):
        if os.path.exists(self.file_path):
            os.remove(self.file_path)
            self._log('Planilha deletada')
    
    def AddOrRemoveRowOrColumn(self, sheet_name, action, index):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb[sheet_name]
        
        if action == 'add_row':
            ws.insert_rows(index)
            self._log(f'Linha {index} adicionada')
        elif action == 'remove_row':
            ws.delete_rows(index)
            self._log(f'Linha {index} removida')
        elif action == 'add_column': 
            ws.insert_cols(index)
            self._log(f'Coluna {index} adicionada')
        elif action == 'remove_column': 
            ws.delete_cols(index)
            self._log(f'Coluna {index} removida')
        
        wb.save(self.file_path)

    def EditCell(self, sheet_name, cell, value):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb[sheet_name]
        ws[cell] = value 
        wb.save(self.file_path)
        self._log(f"Célula {cell} editada para '{value}'")

    def RemoveDuplicates(self, sheet_name):
        df = pd.read_excel(self.file_path, sheet_name = sheet_name)
        df.drop_duplicates(inplace = True)
        with pd.ExcelWriter(self.file_path, engine = 'openpyxl', mode = 'a', if_sheet_exists = 'replace') as writer:
            df.to_excel(writer, sheet_name = sheet_name, index = False)
        self._log(f"Duplicatas removidas da planilha '{sheet_name}'")

    def AutoAdjustColumns(self, sheet_name):
        wb = openpyxl.load_workbook(self.file_path)
        ws = wb[sheet_name]
        
        for col in ws.columns:
            max_length = 0 
            col_letter = openpyxl.utils.get_column_letter(col[0].column)

            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = max_length + 2

        wb.save(self.file_path) 
        self._log(f"Tamanho das colunas ajustado automaticamente na planilha '{sheet_name}'")

    def SortData(self, sheet_name, column, ascending = True):
        df = pd.read_excel(self.file_path, sheet_name = sheet_name)
        df.sort_values(by = column, ascending = ascending, inplace = True)
        with pd.ExcelWriter(self.file_path, engine = 'openpyxl', mode = 'a', if_sheet_exists = 'replace') as writer:
            df.to_excel(writer, sheet_name = sheet_name, index = False)
        self._log(f"Planilha '{sheet_name}' ordenada pela coluna '{column}'")
    
    def CreateChart(self, title, sheet_name, x_col, y_col, chart_type = 'line'):
        df = pd.read_excel(self.file_path, sheet_name = sheet_name)
        plt.figure(figsize = (8,5))
        
        if chart_type == 'line':
            plt.plot(df[x_col], df[y_col], marker = 'o', linestyle = '-')
        elif chart_type == 'bar':
            plt.bar(df[x_col], df[y_col])
        elif chart_type == 'scatter':
            plt.scatter(df[x_col], df[y_col])
        else:
            raise ValueError('[vitalynk.excel] Tipo de gráfico inválido')
    
        plt.xlabel(x_col)
        plt.ylabel(y_col)
        plt.title(title)
        plt.grid(True)
        
        chart_path = f'{sheet_name}_chart.png'
        plt.savefig(chart_path)
        self._log(f"Gráfico '{chart_type}' criado e salvo como '{chart_path}'")
        